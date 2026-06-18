"""
config.py
=========
Centralised configuration / macro store for the Furnace Ranking Optimisation pipeline.

All "macros" in RapidMiner are global key-value variables that every operator can
read or write.  We replicate that here as a plain Python dict (MACROS) plus typed
constant sections for the pipeline parameters that are set once at INPUTS time.

Excel overrides
---------------
Any variable from INPUTS, PIPELINE_MACROS, or the direct MACROS entries can be
overridden from a single Excel sheet (config_overrides.xlsx) with just two columns:

    variable_name   – must exactly match the key name used in Python
    value           – the new value (type is inferred automatically)

The loader searches all three dicts automatically — no need to specify which dict
a variable belongs to.  Variables not listed in the sheet keep their hardcoded
defaults.  Set EXCEL_CONFIG_PATH to None to skip Excel loading entirely.
"""

import os
import logging

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# PATH TO THE EXCEL OVERRIDE FILE
# ---------------------------------------------------------------------------
# Change this to an absolute path or a path relative to this file.
# Set to None to skip Excel loading entirely.
EXCEL_CONFIG_PATH: str | None = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    r"C:\Users\netra.joshi\Documents\POC\Ranking-python-by-rmp\input-data\config_overrides.xlsx",
)

# ---------------------------------------------------------------------------
# HARDCODED DEFAULTS  (used when Excel loading is disabled or the key is absent)
# ---------------------------------------------------------------------------
_INPUTS_DEFAULTS = {
    "fresh_feed_change_set": 0,           # 0 = no change, -1 = force reduction
    "want_to_change_recycle_feed_flow_set": 1,
    "Fur_change_recycle_ethane_limit": 0.3,
    "single_fur_limit": 5,
    "fresh_feed_input": 110,
    "pull_tables_from_db": 0,             # 0 = use local file; 1 = pull from DB
    "NUM_FURNACES": 0,   # ← ADD THIS
    "NUM_PASSES": 0,
}

_PIPELINE_MACROS_DEFAULTS = {
    "split_parameter_name": "feed_type",
    "LBM_initialization_get_pi_data": "active",
    "LBM_preprocessing_inferred_tags_main": "active",
    "ROPT_furnace_coupling": "active",
    "ROPT_external_constraint": "active",
    "ROPT_all_furnace_for_conversion_biasing": "active",
    "ROPT_initialization_use_coilsim": "active",
    "ROPT_use_past_time_output": "active",
    "Optimizer_selector": 6,
    "pass_feed_min_limit": 6.5,
    "pass_step_change": 0.25,
    "max_coke_thickness_limit": 10.0,
    "margin_value_feed_limit": 87.0,
}

# Direct MACROS entries that make sense to override (runtime accumulators excluded).
# The full MACROS dict is built further below; this captures only the overrideable ones.
_MACROS_OVERRIDEABLE = {
    "ranking_cause_indicator": 1,
    "Max_Benefit_SPC": 1000,
    "biasing_condition": 0,
    "mixed_feed_margin": 0,
    "minimum_cracking_furnace_available_check": 0,
    "shc_ratio": 0,
}

# ---------------------------------------------------------------------------
# EXCEL LOADER
# ---------------------------------------------------------------------------

def _coerce(value):
    """
    Coerce a value read from openpyxl to the most natural Python type.
    Whole-number floats (e.g. 6.0) are returned as int.
    """
    if isinstance(value, float) and value == int(value):
        return int(value)
    return value


def _load_excel_overrides(path: str | None) -> dict:
    """
    Read the single 'config_overrides' sheet from *path* and return a flat dict
    of {variable_name: value} for every row that has a recognised key.

    The loader searches _INPUTS_DEFAULTS, _PIPELINE_MACROS_DEFAULTS, and
    _MACROS_OVERRIDEABLE to validate keys — no source_dict column needed.
    Returns an empty dict if path is None or loading fails.
    """
    if path is None:
        return {}

    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not installed – skipping Excel config load.")
        return {}

    if not os.path.exists(path):
        logger.warning("Excel config not found at %s – using hardcoded defaults.", path)
        return {}

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        logger.error("Failed to open Excel config %s: %s – using defaults.", path, exc)
        return {}

    # All valid keys across all three dicts
    all_known_keys = {
        **_INPUTS_DEFAULTS,
        **_PIPELINE_MACROS_DEFAULTS,
        **_MACROS_OVERRIDEABLE,
    }

    sheet_name = wb.sheetnames[0]  # always use the first (and only) sheet
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return {}

    header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    try:
        name_col = header.index("variable_name")
        val_col  = header.index("value")
    except ValueError:
        logger.error("Excel sheet must have 'variable_name' and 'value' columns – using defaults.")
        return {}

    overrides = {}
    for row in rows[1:]:
        if len(row) <= max(name_col, val_col):
            continue
        key   = row[name_col]
        value = row[val_col]
        if key is None:
            continue
        key = str(key).strip()
        if key not in all_known_keys:
            logger.warning("Unknown key '%s' in Excel config – ignored.", key)
            continue
        if value is None:
            logger.warning("Key '%s' has no value in Excel config – keeping default.", key)
            continue
        overrides[key] = _coerce(value)
        logger.debug("Excel override: %s = %r", key, overrides[key])

    return overrides


# ---------------------------------------------------------------------------
# APPLY EXCEL OVERRIDES TO EACH DICT
# ---------------------------------------------------------------------------
_overrides = _load_excel_overrides(EXCEL_CONFIG_PATH)

INPUTS          = {**_INPUTS_DEFAULTS,          **{k: v for k, v in _overrides.items() if k in _INPUTS_DEFAULTS}}
NUM_FURNACES: int = int(INPUTS["NUM_FURNACES"])
NUM_PASSES:   int = int(INPUTS["NUM_PASSES"])
if not logging.root.handlers:
    logging.basicConfig(level=logging.INFO)
logger.info("Pipeline configured for %d furnaces and %d passes per furnace.", NUM_FURNACES, NUM_PASSES)
PIPELINE_MACROS = {**_PIPELINE_MACROS_DEFAULTS, **{k: v for k, v in _overrides.items() if k in _PIPELINE_MACROS_DEFAULTS}}
_macros_overrides = {k: v for k, v in _overrides.items() if k in _MACROS_OVERRIDEABLE}

# ---------------------------------------------------------------------------
# DATABASE / REPOSITORY PATHS  (change to match your environment)
# ---------------------------------------------------------------------------
DB_CONFIG = {
    "repository_entry": r"C:\Users\netra.joshi\Downloads\Ingenero_misc\Charan_Coilsim\Charan_Coilsim\parameterization-ranking-data.xlsx",
    "model_id": 520,
    "tag_prefix": "un.olf%",
    "output_table": "dbo.Furnace_Output",
    "tag_table": "dbo.tag",
}

# ---------------------------------------------------------------------------
# RUNTIME MACRO STORE  (mutable global dict – every module imports this)
# ---------------------------------------------------------------------------
MACROS: dict = {}

MACROS.update(INPUTS)
MACROS.update(PIPELINE_MACROS)

# Additional macros set during INPUTS subprocess (derived):
MACROS.update({
    "end_time": None,

    # Recycle-ethane bounds (derived from Fur_change_recycle_ethane_limit)
    "Fur_change_recycle_ethane_upper_limit": INPUTS["Fur_change_recycle_ethane_limit"],
    "Fur_change_recycle_ethane_lower_limit": -INPUTS["Fur_change_recycle_ethane_limit"],

    # Fresh-feed change flags (one per furnace; all start equal)
    **{f"Fur{i}_Fresh_Feed_Change": INPUTS["fresh_feed_change_set"] for i in range(1, NUM_FURNACES + 1)},

    "Fur_Maximum_Conversion_Single_furnace_limit": INPUTS["single_fur_limit"],
    "Fur_Expected_Fresh_Feed": INPUTS["fresh_feed_input"],
    "Fur_Fresh_Feed_Change": INPUTS["fresh_feed_change_set"],
    "Fur_Want_To_Change_Recycle_Feed_Flow_2": INPUTS["want_to_change_recycle_feed_flow_set"],

    # Deviation / past-time flags (reset per run)
    "deviation_exists": 0,
    "past_time_bypass": 0,

    # Optimiser state flags
    "final_run_optimizer_check": 0,
    "final_run_optimizer_check_init": 0,
    "sum_del_ethylene_final": 0,
    "ranking_cause_indicator": 1,

    # Grid result accumulators
    **{f"Row_{i}_upper_limit_feed": 0 for i in range(1, NUM_FURNACES + 1)},
    **{f"Row_{i}_lower_limit_feed": 0 for i in range(1, NUM_FURNACES + 1)},
    **{f"Row_{i}_step_size_feed": 0 for i in range(1, NUM_FURNACES + 1)},
    **{f"Grid_Row_{i}_conversion_delta": 0 for i in range(1, NUM_FURNACES + 1)},
    **{f"Row_{i}_step_size_conversion": 0 for i in range(1, NUM_FURNACES + 1)},
    **{f"Row_{i}_upper_limit_conversion": 0 for i in range(1, NUM_FURNACES + 1)},
    **{f"Row_{i}_lower_limit_conversion": 0 for i in range(1, NUM_FURNACES + 1)},
    **{f"Row_{i}_Furnace": 0 for i in range(1, NUM_FURNACES + 1)},
    **{f"Row_{i}_part_override": 0 for i in range(1, NUM_FURNACES + 1)},
    **{f"Row_{i}_Feed_flow": 0 for i in range(1, NUM_FURNACES + 1)},
    **{f"Row_{i}_Conversion": 0 for i in range(1, NUM_FURNACES + 1)},
    **{f"Row_{i}_Furnace_condition": "" for i in range(1, NUM_FURNACES + 1)},
    **{f"Row_{i}_Ethylene_Production": 0 for i in range(1, NUM_FURNACES + 1)},
    **{f"Row_{i}_Specific_Energy_consumption": 0 for i in range(1, NUM_FURNACES + 1)},
    **{f"Row_{i}_Current_Recycle_Ethane_Feed": 0 for i in range(1, NUM_FURNACES + 1)},

    # Feed-grid summary macros
    "sum_del_Feed_flow": 0,
    "Min_target_sum_feed_bias": 0,
    "Max_Benefit": 0,
    "Max_Benefit_SPC": 1000,
    "Feed_Grid_Character": "",
    "Conversion_Grid_Success": 0,
    "compare_log_curr_feed_delta": 0,

    # Recycle-ethane grid macros
    "upper_limit_change_in_recycle_ethane": 0,
    "lower_limit_change_in_recycle_ethane": 0,
    "Extra_Recycle_Ethane": 0,

    # Misc optimiser state
    "biasing_condition": 0,
    "mixed_feed_margin": 0,
    "sum_upper_limit_feed": 0,
    "sum_feed_reduction_potential": 0,
    "count_of_good_fur": 0,
    "count_of_no_good_fur": 0,
    "total_fur_available_for_bias": 0,
    "fresh_feed_change": 0,
    "fresh_feed_quantity": 0,
    "shc_ratio": 0,
    "Number_of_rows": 0,
    "extract_value_count": 0,
    "minimum_cracking_furnace_available_check": 0,
    "Fur_Next_Decoking_Furnace": "",
    "min_days_remaining": 0,

    # Past-time timestamp macros
    "24hrs_Timestamp_final_output": None,
    "prev_Timestamp_final_output": None,
    "prev_Timestamp_final_output2": None,

    # Inferred-tag helpers
    "inferred_tags_egs": 0,
    "iteration_inf_tags": 0,

    # Limit macros extracted during pre-processing
    "fuel_gas_pressure_controlvalve_opening_limit": 0,
    "fuel_gas_pressure_limit": 0,
    "damper_opening_limit": 0,
    "quench_ovhd_temp_limit": 0,
    "cgc_suction_pressure_limit": 0,
    "c2_splitter_dp_limit": 0,
    "c2_splitter_btm_c2h4_mol_percent_limit": 0,
    "erc_governor_opening_limit": 0,
    "ethylene_compressure_suction_speed_limit": 0,
    "prc_governor_opening_limit": 0,
    "propylene_compressure_suction_speed_limit": 0,
    "nox_emission_permissible_limit": 0,
    "saturator_drum_pressure_margin_limit": 0,
    "ethane_feed_saturator_drum_overhead_pressure": 0,
    "shc_margin_limit": 0,
    "conversion_bias_threshold_upper_limit": 0,
    "conversion_bias_threshold_lower_limit": 0,
    "conversion_upper_limit_expansion_max_limit": 0,
    "conversion_lower_limit_expansion_max_limit": 0,
    "furnace_step_adjust_feed_grid_limit": 0,

    # Grid result – conversion deltas best found
    **{f"Grid_Row_{i}_conversion_delta_best": 0 for i in range(1, NUM_FURNACES + 1)},
})

# Apply any Excel overrides that target direct MACROS entries
MACROS.update(_macros_overrides)

# ---------------------------------------------------------------------------
# STORE  (replaces RapidMiner's remember/recall mechanism)
# ---------------------------------------------------------------------------
STORE: dict = {}
